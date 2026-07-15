from __future__ import annotations

from io import BytesIO
from typing import Literal

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy.orm import Session

from app.api.deps import require_role
from app.core.database import get_db_session
from app.schemas.common import APISuccess
from app.services.rapports import build_rapport

router = APIRouter(prefix="/rapports")


def _filename(report: dict, extension: str) -> str:
    periode = report.get("periode", {})
    report_type = str(periode.get("type", "rapport")).lower()
    start = str(periode.get("from", "")).replace("-", "")
    end = str(periode.get("to", "")).replace("-", "")
    return f"rapport_{report_type}_{start}_{end}.{extension}"


def _to_rows(report: dict) -> tuple[list[list[str]], list[list[str]], list[list[str]], list[list[str]]]:
    periode = report.get("periode", {})
    kpis = report.get("kpis", {})
    par_terminal = report.get("par_terminal", [])
    serie = report.get("serie_temporelle", [])
    cycles = report.get("cycles", [])

    kpis_rows = [
        ["Champ", "Valeur"],
        ["Type", str(periode.get("type", ""))],
        ["Période", str(periode.get("libelle", ""))],
        ["Du", str(periode.get("from", ""))],
        ["Au", str(periode.get("to", ""))],
        ["Montant total distribué", f"{float(kpis.get('montant_total_distribue', 0.0)):.3f}"],
        ["Nb transactions TR", str(int(kpis.get("nb_transactions_tr", 0)))],
        ["Nb chargements", str(int(kpis.get("nb_chargements", 0)))],
        ["Nb déchargements", str(int(kpis.get("nb_dechargements", 0)))],
        ["Terminaux actifs", str(int(kpis.get("terminaux_actifs", 0)))],
        ["Terminaux inactifs", str(int(kpis.get("terminaux_inactifs", 0)))],
        ["Alertes coffre bas", str(int(kpis.get("nb_alertes_coffre_bas", 0)))],
    ]

    terminal_rows = [[
        "Terminal",
        "Nom",
        "Montant distribué",
        "Nb transactions",
        "Reste coffre dernier",
        "Disponibilité",
        "Alerte coffre bas",
    ]]
    for terminal in par_terminal:
        terminal_rows.append([
            str(terminal.get("terminal_id", "")),
            str(terminal.get("nom", "")),
            f"{float(terminal.get('montant_distribue', 0.0)):.3f}",
            str(int(terminal.get("nb_transactions", 0))),
            f"{float(terminal.get('reste_coffre_dernier', 0.0)):.3f}",
            str(terminal.get("disponibilite", "")),
            "Oui" if terminal.get("alerte_coffre_bas") else "Non",
        ])

    serie_rows = [["Libellé", "Montant distribué"]]
    for item in serie:
        serie_rows.append([
            str(item.get("label", "")),
            f"{float(item.get('montant', 0.0)):.3f}",
        ])

    cycle_rows = [["Terminal", "Nom", "Date déchargement", "Montant charge", "Montant distribué"]]
    for item in cycles:
        cycle_rows.append([
            str(item.get("terminal_id", "")),
            str(item.get("nom", "")),
            str(item.get("datetime_dechargement", "")),
            f"{float(item.get('montant_charge', 0.0)):.3f}",
            f"{float(item.get('montant_distribue', 0.0)):.3f}",
        ])

    return kpis_rows, terminal_rows, serie_rows, cycle_rows


def _build_excel(report: dict) -> BytesIO:
    kpis_rows, terminal_rows, serie_rows, cycle_rows = _to_rows(report)
    workbook = Workbook()
    workbook.remove(workbook.active)

    sheets = [
        ("KPIs", kpis_rows),
        ("Terminaux", terminal_rows),
        ("Série", serie_rows),
        ("Cycles", cycle_rows),
    ]
    header_fill = PatternFill(fill_type="solid", fgColor="E8EFFA")

    for sheet_name, rows in sheets:
        worksheet = workbook.create_sheet(sheet_name)
        for row_index, row_values in enumerate(rows, start=1):
            for col_index, value in enumerate(row_values, start=1):
                cell = worksheet.cell(row=row_index, column=col_index, value=value)
                cell.alignment = Alignment(vertical="top")
                if row_index == 1:
                    cell.font = Font(bold=True)
                    cell.fill = header_fill
            if row_index == 1:
                worksheet.freeze_panes = "A2"
        for column_cells in worksheet.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                cell_value = "" if cell.value is None else str(cell.value)
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            worksheet.column_dimensions[column_letter].width = min(max_length + 4, 42)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def _build_pdf(report: dict) -> BytesIO:
    kpis_rows, terminal_rows, _, _ = _to_rows(report)
    periode = report.get("periode", {})
    buffer = BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=12 * mm,
        leftMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(f"Rapport {periode.get('type', '')}", styles["Title"]))
    elements.append(Paragraph(f"Période: {periode.get('libelle', '')} ({periode.get('from', '')} au {periode.get('to', '')})", styles["BodyText"]))
    elements.append(Spacer(1, 8))

    elements.append(Paragraph("KPIs", styles["Heading2"]))
    elements.append(
        Table(kpis_rows, colWidths=[90 * mm, 70 * mm], repeatRows=1, style=TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e8effa")),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#b8c4d8")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
    )
    elements.append(Spacer(1, 8))

    elements.append(Paragraph("Terminaux", styles["Heading2"]))
    elements.append(
        Table(terminal_rows, repeatRows=1, style=TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e8effa")),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#b8c4d8")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
    )

    document.build(elements)
    buffer.seek(0)
    return buffer


@router.get("", response_model=APISuccess, dependencies=[Depends(require_role("ADMIN", "SUPERVISOR"))])
def get_rapports(
    type: str,
    date: str,
    terminal_id: list[str] | None = Query(default=None),
    db: Session = Depends(get_db_session),
):
    report = build_rapport(db=db, type=type, date_ancre=date, terminal_ids=terminal_id)
    return APISuccess(data=report)


@router.get("/export", dependencies=[Depends(require_role("ADMIN", "SUPERVISOR"))])
def export_rapports(
    type: str,
    date: str,
    format: Literal["pdf", "excel"],
    terminal_id: list[str] | None = Query(default=None),
    db: Session = Depends(get_db_session),
):
    report = build_rapport(db=db, type=type, date_ancre=date, terminal_ids=terminal_id)

    if format == "excel":
        buffer = _build_excel(report)
        headers = {"Content-Disposition": f'attachment; filename="{_filename(report, "xlsx")}"'}
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )

    if format == "pdf":
        buffer = _build_pdf(report)
        headers = {"Content-Disposition": f'attachment; filename="{_filename(report, "pdf")}"'}
        return StreamingResponse(buffer, media_type="application/pdf", headers=headers)

    raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Format d'export invalide")